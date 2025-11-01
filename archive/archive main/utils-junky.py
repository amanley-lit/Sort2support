import os
import json
import re
from datetime import date
import pandas as pd

from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.conf import settings

from .models import Student
from .forms import SignUpForm


from openpyxl import Workbook

# ---------- Load UFLI lessons ----------
def load_ufli_lessons():
    path = os.path.join(settings.BASE_DIR, "main", "static", "main", "data", "ufli_lessons.json")
    with open(path, encoding="utf-8") as f:
        lessons = json.load(f)

    # Ensure every lesson has a total_points field
    for lesson in lessons:
        if "total_points" not in lesson:
            lesson["total_points"] = 5  # fallback

    return lessons

ufli_lessons = load_ufli_lessons()

# ---------- Grouping Logic ----------
def assign_group(student_data, lesson_1, lesson_2):
    # Helpers
    def get_color(score, max_points):
        if score is None or max_points == 0:
            return None
        percent = (score / max_points) * 100
        if percent <= 25: return "Red"
        elif 26 <= percent <= 60: return "Yellow"
        elif 61 <= percent <= 99: return "Green"
        elif percent == 100: return "Blue"
        return None

    def get_instruction_group(score, max_points):
        if score is None: return "Missing"
        if max_points == 3:
            return "Intensive Reteach" if score <= 1 else "Review" if score == 2 else "None"
        elif max_points == 4:
            return "Intensive Reteach" if score <= 1 else "Reteach" if score == 2 else "Review" if score == 3 else "None"
        elif max_points == 5:
            return "Intensive Reteach" if score <= 1 else "Reteach" if score in [2,3] else "Review" if score == 4 else "None"
        elif max_points == 6:
            return "Intensive Reteach" if score <= 1 else "Reteach" if score in [2,3] else "Review" if score in [4,5] else "None"
        return "Unclassified"

    def build_table(groups, concept_name, max_points, missing):
        html = f"<h4>{concept_name} (Max: {max_points})</h4>"
        html += "<table class='assessment-table'><tr><th>Group</th><th>Name</th><th>Score</th></tr>"
        for color in ["Red","Yellow","Green","Blue"]:
            css_class = color.lower() + "-group"
            for name, score in groups[color]:
                html += f"<tr class='{css_class}'><td>{color}</td><td>{name}</td><td>{score}</td></tr>"
        html += "</table>"
        if missing:
            html += f"<p class='note'><em>Not assessed / absent: {', '.join(missing)}</em></p>"
        return html

    def build_weekly_group_table(student_tags, score_key, concept_name, max_points):
        schedule_map = {
            "Intensive Reteach": ["M","Tu","W","Th","F"],
            "Reteach": ["M","W","F"],
            "Review": ["Tu","Th"],
            "None": [],
            "Missing": [],
            "Unclassified": []
        }
        day_order = ["M","Tu","W","Th","F"]
        day_labels = {"M":"M 📘","Tu":"Tu ✏️","W":"W 📚","Th":"Th 🎨","F":"F 🎉"}
        categories = {
            "Intensive Reteach": "🚨 Extra Boost Crew",
            "Reteach": "🔄 Reteach Squad",
            "Review": "🔍 Quick Checkers",
            "None": "🌟 Ready to Fly"
        }
        table_data = {cat:{day:[] for day in day_order} for cat in categories}

        for tag in student_tags:
            group = get_instruction_group(tag[score_key], max_points)
            if group in table_data:
                for day in schedule_map.get(group, []):
                    table_data[group][day].append(tag["name"])

        html = f"<h4>Weekly Group Plan – {concept_name} (Max: {max_points})</h4>"
        html += "<table class='weekly-group-table'>"
        html += "<tr><th>Focus Group</th>" + "".join(f"<th>{day_labels[d]}</th>" for d in day_order) + "</tr>"

        for cat, label in categories.items():
            html += f"<tr><td>{label}</td>"
            for day in day_order:
                names = ", ".join(table_data[cat][day]) or "— No group today —"
                html += f"<td>{names}</td>"
            html += "</tr>"

        html += "</table>"   
        return html

    def build_daily_group_table(student_tags, concept1_name, concept2_name):
        html = "<h4>Daily Grouping Summary</h4>"
        html += "<table class='daily-summary'><tr><th>Name</th>"
        html += f"<th>{concept1_name}</th><th>{concept2_name}</th></tr>"
        for tag in student_tags:
            html += f"<tr><td>{tag['name']}</td>"
            html += f"<td>{tag['concept_1_group']} ({tag['score1']})</td>"
            html += f"<td>{tag['concept_2_group']} ({tag['score2']})</td></tr>"
        html += "</table>"
        return html

    # ---------- Main logic ----------
    max1 = lesson_1["total_points"] if lesson_1 else 5
    max2 = lesson_2["total_points"] if lesson_2 else 5
    concept1_name = lesson_1["concept"] if lesson_1 else "Concept 1"
    concept2_name = lesson_2["concept"] if lesson_2 else "Concept 2"

    concept1_groups = {"Red":[],"Yellow":[],"Green":[],"Blue":[]}
    concept2_groups = {"Red":[],"Yellow":[],"Green":[],"Blue":[]}
    missing1, missing2, student_tags = [], [], []

    for student in student_data:
        name, score1, score2 = student["name"], student["score1"], student["score2"]
        
        student_tags.append({
            "name": name,
            "concept_1_group": group1,
            "concept_2_group": group2,
            "score1": score1,
            "score2": score2,
           
        })
        
        group1 = get_color(score1, max1) or "Missing"
        group2 = get_color(score2, max2) or "Missing"
        student_tags.append({
            "name": name,
            "concept_1_group": group1,
            "concept_2_group": group2,
            "score1": score1,
            "score2": score2,      
        })
        if group1: concept1_groups[group1].append((name,score1))
        else: missing1.append(name)
        if group2: concept2_groups[group2].append((name,score2))
        else: missing2.append(name)

    # ---------- Build HTML ----------
    html = "<div class='grouping-tables'>"
    html += build_table(concept1_groups, concept1_name, max1, missing1)
    html += build_table(concept2_groups, concept2_name, max2, missing2)
    html += build_daily_group_table(student_tags, concept1_name, concept2_name)
    html += "<hr><h3>Weekly Grouping Tables</h3>"
    html += build_weekly_group_table(student_tags, "score1", concept1_name, max1)
    html += build_weekly_group_table(student_tags, "score2", concept2_name, max2)
    html += "</div>"

def safe_sheet_name(name: str) -> str:
    """Make a string safe for Excel sheet names (<=31 chars, no forbidden chars)."""
    cleaned = re.sub(r'[:\\/*?\[\]]', '-', name).strip()
    return cleaned[:31]


def autofit_columns(ws):
    """Resize each column in a worksheet to fit its longest value."""
    for i, col in enumerate(ws.columns, 1):  # enumerate gives you the column index
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def apply_banded_rows(ws, start_row: int = 2):
    """Apply alternating row shading starting from start_row."""
    fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        if row[0].row % 2 == 0:  # even-numbered rows
            for cell in row:
                cell.fill = fill


def style_header_row(ws, row_num: int = 1):
    """Style a header row with bold font, background color, and centered text."""
    header_font = Font(name="Comic Sans MS", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment








def add_group_color_highlighting(ws, start_row=2, last_col="E", group_col="B"):
    """
    Highlight entire rows based on whether the group label in group_col
    contains 'Red', 'Yellow', 'Green', or 'Blue'.
    The range automatically extends to ws.max_row.
    """

    end_row = ws.max_row  # dynamically detect last row with data

    colors = {
        "Red":    "F4CCCC",  # light red
        "Yellow": "FFF2CC",  # light yellow
        "Green":  "D9EAD3",  # light green
        "Blue":   "CFE2F3",  # light blue
    }

    for keyword, hex_color in colors.items():
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        # Formula: look for the keyword anywhere in the group label column
        formula = f'ISNUMBER(SEARCH("{keyword}",${group_col}{start_row}))'
        ws.conditional_formatting.add(
            f"A{start_row}:{last_col}{end_row}",
            FormulaRule(formula=[formula], fill=fill)
        )
@login_required
def edit_uploaded_score(request):
    data = request.session.get("new_entries", [])

    if request.method == "POST":
        formset = StudentFormSet(request.POST)
        if formset.is_valid():
            new_data = []
            for i, form in enumerate(formset):
                name = data[i][0] if i < len(data) else form.cleaned_data.get("name")
                score_1 = form.cleaned_data.get("ufli_score_1")
                score_2 = form.cleaned_data.get("ufli_score_2")
                if name:
                    new_data.append([name, score_1, score_2])
            request.session["new_entries"] = new_data
            data = new_data
    else:
        formset = StudentFormSet(
            initial=[
                {"name": entry[0], "ufli_score_1": entry[1], "ufli_score_2": entry[2]}
                for entry in data
            ]
        )

    for i, form in enumerate(formset):
        if i < len(data) and data[i][0]:
            form.fields["name"].widget.attrs["readonly"] = True

 
    return render(request, "excel_app/upload.html", {"formset": formset, "data": data})

@login_required
def clear_preview_scores(request):
    data = request.session.get("new_entries", [])
    cleared_data = [[entry[0], None, None] for entry in data if entry]
    request.session["new_entries"] = cleared_data
    return redirect("main:upload")


@login_required
def clear_all_students(request):
    Student.objects.filter(teacher=request.user).delete()
    messages.success(request, "✅ All students have been cleared. Ready for a new class!")
    return redirect("main:dashboard")

#@login_required
#def start_new_class(request):
    # Delete all students for this teacher
#    Student.objects.filter(teacher=request.user).delete()

    # Clear session-based preview data
#    request.session["new_entries"] = []
#    request.session["uploaded_students"] = []

#    messages.success(request, "✅ Your roster has been cleared. Ready to start a new class!")
#    return redirect("main:upload")  # or "dashboard" if you prefer

# ---------- Load UFLI lessons ----------
def load_ufli_lessons():
    path = os.path.join(settings.BASE_DIR, "main", "static", "main", "data", "ufli_lessons.json")
    with open(path, encoding="utf-8") as f:
        lessons = json.load(f)

    # Ensure every lesson has a total_points field
    for lesson in lessons:
        if "total_points" not in lesson:
            lesson["total_points"] = 5  # fallback

    return lessons

ufli_lessons = load_ufli_lessons()

def sort2support(request):
    if request.method == "POST":
        # Parse student data
        new_students = []
        i = 1
        while f"student_name_{i}" in request.POST:
            name = request.POST.get(f"student_name_{i}", "").strip()
            score1 = request.POST.get(f"ufli_score_1_{i}") or None
            score2 = request.POST.get(f"ufli_score_2_{i}") or None
            if name:
                new_students.append(Student(
                    name=name,
                    teacher=request.user,
                    ufli_score_1=score1,
                    ufli_score_2=score2
                ))
            i += 1

        Student.objects.filter(teacher=request.user).delete()
        Student.objects.bulk_create(new_students)

        student_data = [{"name": s.name, "score1": s.ufli_score_1, "score2": s.ufli_score_2} for s in new_students]
        lesson_1 = request.session.get("lesson_1")
        lesson_2 = request.session.get("lesson_2")

        html_preview, grouped_data = assign_group(student_data, lesson_1, lesson_2)
        request.session["grouped_data"] = grouped_data

        if request.POST.get("preview_only") == "true":
            return render(request, "excel_app/sort2support.html", {
                "students": new_students,
                "grouping_preview": html_preview,
                "preview_mode": True,
            })
        # Excel export
        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = "Overview"
        current_row = 1

        for concept in grouped_data:
            ws_overview.append([concept["concept_name"], "Group", "Name", "Score"])
            style_header_row(ws_overview, row_num=current_row)
            current_row += 1
            for group in concept["groups"]:
                for s in group["students"]:
                    ws_overview.append([
                        concept["concept_name"],
                        group["group_name"],
                        s["name"],
                        s.get("score", "")
                    ])
                    current_row += 1
                ws_overview.append([])
                current_row += 1
            ws_overview.append([])
            current_row += 1

        autofit_columns(ws_overview)
        add_group_color_highlighting(ws_overview, start_row=2, last_col="D", group_col="B")

        for concept in grouped_data:
            sheet_name = safe_sheet_name(concept["concept_name"])
            ws = wb.create_sheet(title=sheet_name)
            ws.merge_cells("A1:C1")
            ws["A1"] = concept["concept_name"]
            ws.append(["Group", "Name", "Score"])
            style_header_row(ws, row_num=2)
            for group in concept["groups"]:
                for s in group["students"]:
                    ws.append([group["group_name"], s["name"], s.get("score", "")])
                ws.append([])
            autofit_columns(ws)
            add_group_color_highlighting(ws, start_row=3, last_col="C", group_col="A")

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="grouped_roster.xlsx"'
        wb.save(response)
        return response

    # GET fallback
    students = Student.objects.filter(teacher=request.user)
    return render(request, "excel_app/sort2support.html", {
        "students": students,
        "preview_mode": False,
    })

