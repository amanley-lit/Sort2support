from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from main.utils.export_excel import generate_excel
from main.models import Student, Roster
from .forms import SignUpForm, AddStudentForm
from main.main_utils import (
    load_ufli_lessons,
    assign_group,
    get_instruction_group,
    get_color_class,
)
from datetime import datetime
import io, re, pandas as pd, openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ---------- Auth Views ----------

def home(request):
    """Renders the public homepage."""
    return render(request, "main/home.html")

def signup(request):
    """Handles user registration and redirects to dashboard on success."""
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect("main:dashboard")
    else:
        form = SignUpForm()
    return render(request, "main/signup.html", {"form": form})

def login_view(request):
    """Authenticates user and redirects to dashboard."""
    form = AuthenticationForm(data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.get_user()
        login(request, user)
        return redirect("main:dashboard")
    return render(request, "main/login.html", {"form": form})

def logout_view(request):
    """Logs out the current user and redirects to homepage."""
    logout(request)
    return redirect("main:home")

# ---------- Dashboard ----------

@login_required
def dashboard(request):
    """Displays student data, grouping logic, and lesson metadata."""

    context = {}  # ✅ Define early so it's safe to use below

    # Add static context
    ufli_lessons = load_ufli_lessons()
    context["ufli_lessons"] = ufli_lessons
    context["saved_rosters"] = Roster.objects.filter(user=request.user)
    
    # Restore Step 1 selections from session
    
    if "lesson_1_id" in request.session and "lesson_2_id" in request.session:
        context["lesson_1_id"] = request.session["lesson_1_id"]
        context["lesson_2_id"] = request.session["lesson_2_id"]
        context["lesson_1_name"] = request.session.get("lesson_1_name")
        context["lesson_2_name"] = request.session.get("lesson_2_name")
        context["lesson_1_max"] = request.session.get("lesson_1_max")
        context["lesson_2_max"] = request.session.get("lesson_2_max")
        context["preview_data"] = request.session.get("preview_data", [])
        context["student_count"] = request.session.get("student_count", 0)
        context["entry_mode"] = request.session.get("entry_mode", "")
        context["groups"] = request.session.get("groups", [])

    # --- Initialize defaults ---

    students = Student.objects.filter(teacher=request.user)
    student_tags = {s.name: s.tag for s in students}
    context["student_tags"] = student_tags

    # --- Handle POST actions ---
    if request.method == "POST":

        # Step 1: Save selected lessons (local message, no redirect)
        if "save_lessons" in request.POST:
            lesson_1_id = request.POST.get("lesson_1")
            lesson_2_id = request.POST.get("lesson_2")

            lesson_1 = next((l for l in ufli_lessons if str(l.get("number")) == str(lesson_1_id)), None)
            lesson_2 = next((l for l in ufli_lessons if str(l.get("number")) == str(lesson_2_id)), None)

            if not lesson_1 or not lesson_2:
                context["step1_error"] = "❌ Could not find one or both selected lessons."

            else:
                request.session["step1_done"] = True
                request.session["lesson_1_id"] = str(lesson_1["number"])
                request.session["lesson_2_id"] = str(lesson_2["number"])
                request.session["lesson_1_name"] = lesson_1["concept"]
                request.session["lesson_2_name"] = lesson_2["concept"]
                request.session["lesson_1_max"] = lesson_1["total_points"]
                request.session["lesson_2_max"] = lesson_2["total_points"]
                request.session["lesson_meta"] = {
                    "lesson_1": {
                        "id": str(lesson_1["number"]), 
                        "name": lesson_1["concept"], 
                        "max": lesson_1["total_points"], 
                        "full": lesson_1,
                    },
                    "lesson_2": {
                        "id": str(lesson_2["number"]), 
                        "name": lesson_2["concept"], 
                        "max": lesson_2["total_points"], 
                        "full": lesson_2
                    },
                }
                context["lesson_1_id"] = str(lesson_1["number"])
                context["lesson_2_id"] = str(lesson_2["number"])

                # ✅ Only set success message if both lessons are valid
                if lesson_1 and lesson_2:
                    context["step1_success"] = f"✅ Lessons saved: {lesson_1['concept']} and {lesson_2['concept']}."


            # Rehydrate context just like in your dashboard view
            context.update({
                "step1_done": request.session.get("step1_done", False),
                "ufli_lessons": ufli_lessons,
                "lesson_1_id": request.session.get("lesson_1_id"),
                "lesson_2_id": request.session.get("lesson_2_id"),
                "lesson_1": lesson_1,
                "lesson_2": lesson_2,
                "preview_data": request.session.get("preview_data", []),
                "entry_mode": request.session.get("entry_mode", "paste"),
                "grouped_html": request.session.get("grouped_html"),
                "grouped_data": request.session.get("grouped_data"),
                "just_grouped": request.session.pop("just_grouped", False),
                "students": students,
                "student_tags": student_tags,
            })
            return render(request, "main/dashboard.html", context)


        # Step 2: Choose entry mode
        elif "entry_mode" in request.POST:
            request.session["entry_mode"] = request.POST["entry_mode"]
            return redirect("main:dashboard")

        # Step 2a: Paste roster
        elif "process_roster_raw" in request.POST:
            raw_text = request.POST.get("roster_raw", "")
            lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
            preview_data = [{"name": name, "score1": 0, "score2": 0} for name in lines]

            if not preview_data:
                messages.error(request, "❌ Paste failed: no names found.", extra_tags="step2")
            else:
                request.session["preview_data"] = preview_data
                request.session["entry_mode"] = "preview"
                request.session["student_count"] = len(preview_data)
                request.session["step2_done"] = True
                request.session["step3_open"] = True

                messages.success(request, f"✅ {len(preview_data)} names processed from paste.", extra_tags="step2")
  
            request.session["step2_done"] = True
            return redirect("main:dashboard")

        # Step 2b: Upload roster
        elif "process_roster_upload" in request.POST and request.FILES.get("roster_file"):
            try:
                df = pd.read_excel(request.FILES["roster_file"])
                df.columns = [str(c).strip().lower().replace(" ", "") for c in df.columns]
                if not {"name", "score1", "score2"}.issubset(df.columns):
                    messages.error(request, "❌ Missing required columns: name, score1, score2.")
                    return redirect("main:dashboard")

                preview_data = []
                for _, row in df.iterrows():
                    name = str(row.get("name", "")).strip()
                    score1 = row.get("score1")
                    score2 = row.get("score2")
                    missing_score1 = score1 in (None, "", "nan")
                    missing_score2 = score2 in (None, "", "nan")
                    try:
                        score1 = int(score1) if not missing_score1 else 0
                    except (ValueError, TypeError):
                        score1, missing_score1 = 0, True
                    try:
                        score2 = int(score2) if not missing_score2 else 0
                    except (ValueError, TypeError):
                        score2, missing_score2 = 0, True
                    if name:
                        preview_data.append({
                            "name": name,
                            "score1": score1,
                            "score2": score2,
                            "missing_score1": missing_score1,
                            "missing_score2": missing_score2,
                        })
   

                if not preview_data:
                    messages.error(request, "❌ Upload failed: no valid rows.", extra_tags="step2")
                else:
                    request.session["preview_data"] = preview_data
                    request.session["student_count"] = len(preview_data)
                    request.session["entry_mode"] = "upload"
                    request.session["step2_done"] = True
                    request.session["step3_open"] = True 
                    messages.success(request, f"✅ {len(preview_data)} names processed from upload.", extra_tags="step2")

            except Exception:
                messages.error(request, "❌ Upload failed: invalid file.", extra_tags="step2")

            return redirect("main:dashboard")


        # Step 2c: Load / delete rosters

        elif "load_selected_roster" in request.POST:
            roster_id = request.POST.get("roster_id")
            try:
                roster = Roster.objects.get(id=roster_id, user=request.user)
                request.session["preview_data"] = roster.data
                request.session["student_count"] = len(roster.data)
                request.session["step2_done"] = True
                request.session["step3_open"] = True 
                request.session["entry_mode"] = "load"
                request.session["loaded_roster_name"] = roster.name

                messages.success(request, f"✅ Roster '{roster.name}' loaded.", extra_tags="step2c")
            except Roster.DoesNotExist:
                messages.error(request, "❌ Selected roster not found.", extra_tags="step2c")
            return redirect("main:dashboard")

        elif "delete_roster" in request.POST:
            roster_id = request.POST.get("roster_id")
            try:
                roster = Roster.objects.get(id=roster_id, user=request.user)
                roster.delete()
                messages.success(request, f"🗑️ Roster '{roster.name}' deleted.", extra_tags="step2c")
            except Roster.DoesNotExist:
                messages.error(request, "❌ Could not delete roster.", extra_tags="step2c")
            return redirect("main:dashboard")


        # Step 3: Save scores
        elif "save_roster_raw" in request.POST:
            student_count = request.session.get("student_count", 0)
            preview_data = []

            # Parse existing student rows
            for i in range(1, student_count + 1):
                #skip deleted rows
                if request.POST.get(f"delete_{i}"):
                    continue
                name = request.POST.get(f"name_{i}", "").strip()
                score1 = request.POST.get(f"score1_{i}", "")
                score2 = request.POST.get(f"score2_{i}", "")
                if name:
                    try:
                        score1 = int(score1) if score1 else 0
                    except (ValueError, TypeError):
                        score1 = 0
                    try:
                        score2 = int(score2) if score2 else 0
                    except (ValueError, TypeError):
                        score2 = 0
                    preview_data.append({"name": name, "score1": score1, "score2": score2})

            # Handle new student row
            new_name = request.POST.get("new_name", "").strip()
            if new_name:
                new_score1 = request.POST.get("new_score1") or 0
                new_score2 = request.POST.get("new_score2") or 0
                try:
                    new_score1 = int(new_score1)
                except (ValueError, TypeError):
                    new_score1 = 0
                try:
                    new_score2 = int(new_score2)
                except (ValueError, TypeError):
                    new_score2 = 0

                preview_data.append({"name": new_name, "score1": new_score1, "score2": new_score2})

            roster_name = request.POST.get("roster_name", "").strip()
            
            if not roster_name:
                messages.error(request, "❌ Roster name is required.")
                return redirect("main:dashboard")

            roster = Roster.objects.create(name=roster_name, user=request.user)

            if not preview_data:
                context["upload_error"] = "❌ No valid student names found. Roster not saved."
            elif not roster_name:
                context["upload_error"] = "❌ Roster name is required."
            else:
                Roster.objects.update_or_create(
                    user=request.user,
                    name=roster_name,
                    defaults={"data": preview_data}
                )

                # ✅ Mark Step 3 complete and unlock Step 4
                
                context["roster_uploaded"] = True
                context["roster_name"] = roster_name
                request.session["preview_data"] = preview_data
                request.session["student_count"] = len(preview_data)
                request.session["loaded_roster_name"] = roster.name
                request.session["step3_done"] = True  # ✅ Add this inside the final else block
                request.session["step4_open"] = True

                messages.success(
                    request,
                    f"✅ Roster '{roster_name}' saved with {len(preview_data)} students.",
                    extra_tags="step3"
                )

            # Update context for rendering

            context.update({
                "step1_done": request.session.get("step1_done", False),
                "ufli_lessons": ufli_lessons,
                "lesson_1_id": request.session.get("lesson_1_id"),
                "lesson_2_id": request.session.get("lesson_2_id"),
                "lesson_1": next((l for l in ufli_lessons if str(l.get("number")) == str(request.session.get("lesson_1_id"))), None),
                "lesson_2": next((l for l in ufli_lessons if str(l.get("number")) == str(request.session.get("lesson_2_id"))), None),
                "preview_data": preview_data,
                "entry_mode": request.session.get("entry_mode", "paste"),
                "grouped_html": request.session.get("grouped_html"),
                "grouped_data": request.session.get("grouped_data"),
                "just_grouped": request.session.pop("just_grouped", False),
                "students": students,
                "student_tags": student_tags,
                "loaded_roster_name": request.session.get("loaded_roster_name"),
                "step3_done": request.session.get("step3_done", False),
                "step4_open": request.session.get("step4_open", False),
            })


            messages.success(
                request,
                f"✅ Roster '{roster_name}' saved with {len(preview_data)} students.",
                extra_tags="step3"
            )
            return render(request, "main/dashboard.html", context)


        # Step 4: Sort2Support
        elif "sort2support" in request.POST:
            lesson_1_id = request.POST.get("lesson_1")
            lesson_2_id = request.POST.get("lesson_2")
            lesson_1 = next((l for l in ufli_lessons if str(l.get("number")) == lesson_1_id), None)
            lesson_2 = next((l for l in ufli_lessons if str(l.get("number")) == lesson_2_id), None)
            preview_data = request.session.get("preview_data", [])

            if not lesson_1 or not lesson_2 or not preview_data:
                context["group_error"] = "❌ Missing data for export. Please group students with Sort2Support in Step 4 first."
                return render(request, "main/dashboard.html", context)

            grouped_html, grouped_data = assign_group(preview_data, lesson_1, lesson_2, student_tags)

            request.session["grouped_daily"] = grouped_data["daily"]
            request.session["grouped_data"] = grouped_data
            request.session["grouped_html"] = grouped_html
            request.session["just_grouped"] = True

            request.session["lesson_meta"] = {
                "lesson_1": {
                    "id": str(lesson_1["number"]),
                    "name": lesson_1["concept"],
                    "max": lesson_1["total_points"],
                    "full": lesson_1,
                },
                "lesson_2": {
                    "id": str(lesson_2["number"]),
                    "name": lesson_2["concept"],
                    "max": lesson_2["total_points"],
                    "full": lesson_2,
                }
            }
            messages.success(request, "✅ Students grouped successfully.", extra_tags="step4")

            context["group_success"] = f"✅ Students grouped for {lesson_1['concept']} and {lesson_2['concept']}."
            context.update({
                "step1_done": request.session.get("step1_done", False),
                "ufli_lessons": ufli_lessons,
                "lesson_1": lesson_1,
                "lesson_2": lesson_2,
                "lesson_1_id": lesson_1_id,
                "lesson_2_id": lesson_2_id,
                "preview_data": preview_data,
                "entry_mode": request.session.get("entry_mode", "paste"),
                "grouped_html": grouped_html,
                "grouped_data": grouped_data,
                "just_grouped": True,
                "students": students,
                "student_tags": student_tags,
            })


            return render(request, "main/dashboard.html", context)

        # Step 5: Finalize grouped data

        elif "finalize_groups" in request.POST:
            grouped_data = request.session.get("grouped_data")
            grouped_html = request.session.get("grouped_html")
            lesson_meta = request.session.get("lesson_meta", {})
            preview_data = request.session.get("preview_data", [])

            if not grouped_data or not grouped_html or not lesson_meta:
                messages.error(request, "❌ Grouped data missing. Please complete Step 4 first.", extra_tags="step5")
                return render(request, "main/dashboard.html", context)

            # Mark Step 5 complete
            request.session["step5_done"] = True
            request.session["step5_open"] = True
            request.session["just_finalized"] = True

            messages.success(request, "✅ Groups finalized and ready for export.", extra_tags="step5")

            context.update({
                "grouped_data": grouped_data,
                "grouped_html": grouped_html,
                "lesson_meta": lesson_meta,
                "preview_data": preview_data,
                "step5_done": True,
                "step5_open": True,
                "just_finalized": True,
            })

            return render(request, "main/dashboard.html", context)
            

    # --- After POST handling, hydrate context for GET render ---
    lesson_1_id = request.session.get("lesson_1_id")
    lesson_2_id = request.session.get("lesson_2_id")
    lesson_1 = next((l for l in ufli_lessons if str(l.get("number")) == str(lesson_1_id)), None)
    lesson_2 = next((l for l in ufli_lessons if str(l.get("number")) == str(lesson_2_id)), None)

    grouped_data = request.session.get("grouped_data")
    grouped_html = request.session.get("grouped_html")
    lesson_meta = request.session.get("lesson_meta", {})
    just_grouped = request.session.pop("just_grouped", False)
    just_finalized = request.session.pop("just_finalized", False)

    entry_mode = request.POST.get("entry_mode") or request.session.get("entry_mode", "paste")
    request.session["entry_mode"] = entry_mode
    preview_data = request.session.get("preview_data", [])

    # --- Validate grouped_data ---
    if not isinstance(grouped_data, dict):
        print(f"grouped_data is None or invalid: {grouped_data}")
        grouped_data = {}

    # --- Step tracking ---  
    step1_done = request.session.get("step1_done", False)
    step2_done = request.session.get("step2_done", False)
    step3_done = request.session.get("step3_done", False)
    step4_done = request.session.get("step4_done", False)
    step5_done = request.session.get("step5_done", False)
    step5_open = request.session.get("step5_open", False)

    # --- Determine current step ---

    if not step1_done:
        current_step = "step1"
    elif not step2_done:
        current_step = "step2"
    elif not step3_done:
        current_step = "step3"
    elif not step4_done:
        current_step = "step4"
    elif not step5_done:
        current_step = "step5"
    else:
        current_step = "complete"


    # --- Build final context ---

    context.update({
        "groups1": grouped_data.get("concept1", {}),
        "groups2": grouped_data.get("concept2", {}),
        "concept1_name": lesson_1["name"] if lesson_1 and "name" in lesson_1 else "Concept 1",
        "concept2_name": lesson_2["name"] if lesson_2 and "name" in lesson_2 else "Concept 2",
        "ufli_lessons": ufli_lessons,
        "lesson_1": lesson_1,
        "lesson_2": lesson_2,
        "lesson_1_id": lesson_1_id,
        "lesson_2_id": lesson_2_id,
        "preview_data": preview_data,
        "entry_mode": entry_mode,
        "grouped_html": grouped_html,
        "grouped_data": grouped_data,
        "just_grouped": just_grouped,
        "students": students,
        "student_tags": student_tags,
        "step1_done": step1_done,
        "step2_done": step2_done,
        "step3_done": step3_done,
        "step4_done": step4_done,
        "step5_done": step5_done,
        "step5_open": step5_open,
        "current_step": current_step,
    })


   
    # Optional: saved rosters list
    # context["saved_rosters"] = Roster.objects.filter(user=request.user).order_by("-created_at")

    # Clear any one-time flags
    request.session.pop("step2_open", None)

    return render(request, "main/dashboard.html", context)


# ---------- Upload / Reset ----------
@login_required
def upload_page(request):
    if request.method == "POST" and request.FILES.get("file"):
        wb = openpyxl.load_workbook(request.FILES["file"])
        sheet = wb.active

        preview_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, score1, score2 = row[:3]
            if name:
                preview_data.append({
                    "name": str(name).strip(),
                    "score1": int(score1) if score1 else 0,
                    "score2": int(score2) if score2 else 0,

                })

        request.session["preview_data"] = preview_data
        request.session["entry_mode"] = "preview"
        return redirect("main:dashboard")

    return render(request, "main/dashboard.html")

def download_template(request):
    """
    Generate and return a blank Excel roster template with styled headers.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster Template"

    # Define header labels
    headers = ["Name", "Score 1", "Score 2"]
    ws.append(headers)

    # Define styles: blue fill, white bold font, centered
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Apply styles to header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add an example row
    example_row = ["Alice S.", "2", "3"]
    ws.append(example_row)

    # Style the example row (italic, gray text)
    example_font = Font(color="808080", italic=True)
    for cell in ws[2]:
        cell.font = example_font



    # Optionally set column widths for readability
    column_widths = [15, 15, 15, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = width

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="roster_template.xlsx"'
    return response

@login_required
def load_previous_roster(request):
    last_roster = Roster.objects.filter(user=request.user).order_by("-created_at").first()
    if last_roster:
        request.session["preview_data"] = last_roster.data
        request.session["student_count"] = len(last_roster.data)
        request.session["just_loaded"] = True
    return redirect("main:dashboard")  

@login_required
def reset_saved_scores(request):
    """Clears all student scores but preserves names, and clears preview/grouped session data."""
    students = Student.objects.filter(teacher=request.user)
    for s in students:
        s.ufli_score_1 = None
        s.ufli_score_2 = None
        s.save()

    # Clear any preview/grouped session data so the dashboard refreshes cleanly
    for key in ["preview_data", "grouped_html", "grouped_data"]:
        request.session.pop(key, None)

    messages.success(request, "🧹 All scores cleared, but student names remain.")
    return redirect("main:dashboard")

@login_required
def reset_class(request):
    """Deletes all students for the current teacher and clears preview/grouped session data."""
    Student.objects.filter(teacher=request.user).delete()

    # Clear any preview/grouped session data so the dashboard refreshes cleanly
    
    for key in ["preview_data", "grouped_html", "grouped_data"]:
        request.session.pop(key, None)
        request.session.pop("grouped_data", None)
        request.session.pop("grouped_html", None)
        request.session.pop("lesson_meta", None)
        request.session.pop("step4_done", None)
        request.session.pop("step5_done", None)
        request.session.pop("step4_open", None)
        request.session.pop("step5_open", None)

    messages.success(request, "🔄 Entire class reset successfully.")
    return redirect("main:dashboard")

# ---------- Export Logic ----------

def sheet_name_with_date(title: str) -> str:
    """
    Sanitize a string for use as an Excel sheet name, always appending today's date.
    - Removes invalid Excel characters
    - Appends YYYY-MM-DD
    - Truncates to 31 characters (Excel limit)
    """
    # Remove invalid Excel characters
    safe = re.sub(r'[:\\/*?\[\]]', '-', title).strip()

    # Append today's date
    today_str = datetime.now().strftime("%Y-%m-%d")
    safe = f"{safe} {today_str}"

    # Excel sheet names max length = 31
    if len(safe) > 31:
        safe = safe[:28] + "..."

    return safe

def autofit_columns(ws):
    """Resize each column in a worksheet to fit its longest value."""
    for i, col in enumerate(ws.columns, 1):  # i is the column index
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

def style_header_row(ws, row_num: int = 1):
    """Style a header row with bold font, background color, and centered text."""
    header_font = Font(name="Comic Sans MS", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def get_fill(score):
    if score is None:
        return None
    if score <= 25:
        return PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # red
    elif score <= 60:
        return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
    elif score <= 99:
        return PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # green
    else:  # 100%
        return PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")  # blue
    

# ---------- Export ----------
@login_required
def export_grouped_excel(request):

    # Pull lesson metadata from session
    lesson_1_name = request.session.get("lesson_1_name", "Concept 1")
    lesson_2_name = request.session.get("lesson_2_name", "Concept 2")

    print("✅ Export view triggered")
    
    # grouped_data = request.session.get("grouped_data")

    daily_data = request.session.get("grouped_daily", [])
    grouped_data = request.session.get("grouped_data", {})

    if not grouped_data or not daily_data:
        messages.error(
            request,
            "❌ Missing data for export. Please group students with Sort2Support in Step 4 first."
        )
        return redirect("main:dashboard")

    print("✅ lesson_1_name:", lesson_1_name)
    print("✅ lesson_2_name:", lesson_2_name)
    print("✅ daily_data:", list(grouped_data.keys()) if grouped_data else "❌ grouped_data missing")


    if not grouped_data or not daily_data:
        messages.error(
            request, 
            "❌ Missing data for export. Please click Sort2Support first."
        )
        return redirect("main:dashboard")

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)   # Remove default sheet

    print("✅ Workbook built with sheets:", wb.sheetnames)

    # ✅ Sheet 1: Daily Summary
    summary = wb.create_sheet(title="Assessment: Grouping Assignments")
    summary.append([
        "Student", 
        f"Group 1 ({lesson_1_name})", 
        f"Group 2 ({lesson_2_name})"
    ])

    for student in daily_data:
        summary.append([
            student["name"],
            student["group_1"],
            student["group_2"]
        ])
    
    summary.freeze_panes = "A2"
    autofit_columns(summary)


    # --- Sheet 2: Concept 1 Groups ---
    concept1_sheet = wb.create_sheet(title=sheet_name_with_date(lesson_1_name))
    concept1_sheet.append(["Group", "Name", "Score"])
    style_header_row(concept1_sheet)

    for group_name, students in grouped_data.get("concept1", {}).items():
        for name, score in students:
            concept1_sheet.append([group_name, name, score])

    concept1_sheet.freeze_panes = "A2"
    autofit_columns(concept1_sheet)
    add_group_color_highlighting(concept1_sheet)



    # ✅ Sheet 3: Concept 2 Groups
    concept2_sheet = wb.create_sheet(title=sheet_name_with_date(lesson_2_name))
    concept2_sheet.append(["Group", "Name", "Score"])
    style_header_row(concept2_sheet)

    for group_name, students in grouped_data.get("concept2", {}).items():
        for name, score in students:
            concept2_sheet.append([group_name, name, score])

    concept2_sheet.freeze_panes = "A2"
    autofit_columns(concept2_sheet)
    add_group_color_highlighting(concept2_sheet)


    # ✅ Sheet 4: Weekly Plan
    weekly_sheet = wb.create_sheet(title="Weekly Plan")
    weekly_sheet.append(["Focus Group", "Tu", "W", "Th", "F"])
    style_header_row(weekly_sheet)

    # Build table_data
    day_order = ["Tu", "W", "Th", "F"]
    schedule_map = {
        "Intensive Reteach": ["M", "Tu", "W", "F"],
        "Reteach": ["M", "W", "F"],
        "Review": ["Tu", "Th"],
        "None": [],
        "Missing": [],
        "Unclassified": []
    }
    categories = {
        "Intensive Reteach": "🚨 Extra Boost Crew",
        "Reteach": "🔄 Reteach Squad",
        "Review": "🔍 Quick Checkers",
        "None": "🌟 Ready to Fly"
    }

    table_data = {cat: {day: [] for day in day_order} for cat in categories}
    score_key = "score1"  # or "score2" depending on context
    max_points = request.session.get("lesson_1_max", 5)

    for tag in grouped_data.get("tags", []):
        name = tag.get("name")
        score = tag.get(score_key)
        if name is None or score is None:
            continue
        group = get_instruction_group(score, max_points)
        for day in schedule_map.get(group, []):
            if day in table_data[group]:
                table_data[group][day].append(name)

    # Append rows
    for group_key, label in categories.items():
        row = [label]
        for day in day_order:
            names = ", ".join(table_data[group_key][day]) or "No group today—"
            row.append(names)
        weekly_sheet.append(row)

    autofit_columns(weekly_sheet)


def add_group_color_highlighting(ws, start_row=2, last_col="C", group_col="A"):
    """
    Highlight entire rows based on whether the group label in group_col
    contains 'Red', 'Yellow', 'Green', or 'Blue'.
    The range automatically extends to ws.max_row.
    """

    end_row = ws.max_row  # dynamically detect last row with data
    
    ws.freeze_panes = "A2"

    colors = {
        "Red":    "F4CCCC",  # light red
        "Yellow": "FFF2CC",  # light yellow
        "Green":  "D9EAD3",  # light green
        "Blue":   "CFE2F3",  # light blue
    }

    for keyword, hex_color in colors.items():
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        # Formula: look for the keyword anywhere in the group label column
        formula = f'ISNUMBER(SEARCH("{keyword}",${group_col} & ROW()))'
        ws.conditional_formatting.add(
            f"A{start_row}:{last_col}{end_row}",
            FormulaRule(formula=[formula], fill=fill)
        )


#        sheet_title = lesson_names.get(concept_key, concept_key)
#        ws = wb.create_sheet(title=sheet_title)
#        ws.append(["Group", "Name", "Score"])

#        for group_name, students in concept_groups.items():
#            for name, score in students:
#                ws.append([group_name, name, score])

#        ws.freeze_panes = "A2"

    # ✅ Return workbook as response
#    response = HttpResponse(
#        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#    )
#    filename = f"Sort2Support_Groups_{datetime.now():%Y%m%d}.xlsx"
#    response["Content-Disposition"] = f'attachment; filename="{filename}"'
#    wb.save(response)

#    return response




# ---------- Excel Templates  ----------
@login_required
def update_scores(request):
    """Updates student scores from dashboard form submission."""
    if request.method == "POST":
        students = Student.objects.filter(teacher=request.user)
        for student in students:
            score_1 = request.POST.get(f"score_1_{student.id}", None)
            score_2 = request.POST.get(f"score_2_{student.id}", None)

            score_1 = score_1 if score_1 != "" else None
            score_2 = score_2 if score_2 != "" else None

            if score_1 is not None:
                try:
                    student.ufli_score_1 = int(score_1)
                except ValueError:
                    student.ufli_score_1 = None
            if score_2 is not None:
                try:
                    student.ufli_score_2 = int(score_2)
                except ValueError:
                    student.ufli_score_2 = None

            student.save()

        messages.success(request, "✅ Scores updated successfully.")
    return redirect("main:dashboard")



#    ws.append(["Name", "Score 1", "Score 2"])
#    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#    response["Content-Disposition"] = 'attachment; filename="Sort2Support_Template.xlsx"'
#    wb.save(response)
#    return response

@login_required
def generate_excel_view(request):
    grouped_data = request.session.get("grouped_data", {})
    lesson_meta = request.session.get("lesson_meta", {})


    lesson_1 = lesson_meta.get("lesson_1", {})
    lesson_2 = lesson_meta.get("lesson_2", {})

    print("✅ lesson_1:", lesson_1)
    print("✅ lesson_2:", lesson_2)
    print("✅ grouped_data keys:", list(grouped_data.keys()) if grouped_data else "None")

    if not grouped_data or not lesson_1 or not lesson_2:
        messages.error(request, "Missing data for export. Please click Sort2Support first.")
        return redirect("main:dashboard")

    concept1_name = lesson_1.get("name")
    concept2_name = lesson_2.get("name")
    max1 = lesson_1.get("max")
    max2 = lesson_2.get("max")

    print("✅ grouped_data keys:", list(grouped_data.keys()))

    # Generate workbook
    wb = generate_excel(grouped_data, lesson_1, lesson_2)  # must return a Workbook

    # Write workbook into a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return as downloadable response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_export.xlsx"'
    return response

def generate_excel(grouped_data, lesson_1, lesson_2):
    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    day_order = ["M", "Tu", "W", "Th", "F"]
    day_labels = {"M":"M","Tu":"Tu","W":"W","Th":"Th","F":"F"}
    schedule_map = {
        "Intensive Reteach": ["M", "Tu", "W", "F"],
        "Reteach": ["M", "W", "F"],
        "Review": ["Tu", "Th"],
        "None": [],
        "Missing": [],
        "Unclassified": []
    }
    categories = {
        "Intensive Reteach": "🚨 Extra Boost Crew",
        "Reteach": "🔄 Reteach Squad",
        "Review": "🔍 Quick Checkers",
        "None": "🌟 Ready to Fly"
    }
    group_colors = {
        "Extra Boost Crew": "FFC0CB",  # Pink
        "Reteach Squad": "FFFACD",     # Lemon
        "Quick Checkers": "CCFFCC",    # Light green
        "Ready to Fly": "ADD8E6",      # Light blue
    }

    lesson_names = {
        "concept1": lesson_1["name"],
        "concept2": lesson_2["name"]
    }

    for concept_key, groups in grouped_data.items():
        if concept_key == "tags":
            continue

        sheet_title = sheet_name_with_date(lesson_names.get(concept_key, concept_key))
        ws = wb.create_sheet(title=sheet_title)

        if concept_key == "daily":
            ws.append(["Student", "Group 1", "Concept 1", "Group 2", "Concept 2"])
            for student in groups:
                name = student.get("name", "")
                group_1 = student.get("group_1", "")
                concept_1 = student.get("concept_1", "")
                group_2 = student.get("group_2", "")
                concept_2 = student.get("concept_2", "")
                ws.append([name, group_1, concept_1, group_2, concept_2])

        elif isinstance(groups, dict):
            ws.append(["Group", "Student", "Score"])
            score_key = "score1" if concept_key == "concept1" else "score2"

            for group_name, students in groups.items():
                for student in students:
                    # ✅ Handle dicts
                    if isinstance(student, dict):
                        name = student.get("name", "")
                        score = student.get(score_key, "")
                    # ✅ Handle 2-element tuples/lists
                    elif isinstance(student, (list, tuple)) and len(student) == 2:
                        name, score = student
                    # ✅ Handle plain strings
                    elif isinstance(student, str):
                        name = student
                        score = ""
                    # ✅ Handle unexpected formats
                    else:
                        print(f"⚠️ Unexpected student format in {group_name}:", student)
                        continue

                    # ✅ Flatten deeply nested values
                    if isinstance(name, (list, tuple)):
                        name = ", ".join(str(x) for x in name)
                    if isinstance(score, (list, tuple)):
                        score = ", ".join(str(x) for x in score)

                    ws.append([group_name, name, score])





        else:
            print(f"⚠️ Skipping {concept_key} — unexpected structure:", type(groups))
            continue

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        print("🔍 concept_key:", concept_key)
        print("🔍 groups type:", type(groups))

        if isinstance(groups, list):
            print("🔍 groups sample:", groups[:1])
        elif isinstance(groups, dict):
            print("🔍 groups sample:", list(groups.items())[:1])
        else:
            print("🔍 groups sample:", groups)

        ws.freeze_panes = "A2"
        add_group_color_highlighting(ws, start_row=2, last_col="C", group_col="A")

        # Leave a blank row before weekly plan
        ws.append([])

        # Weekly Plan Header
        ws.append(["Focus Group"] + [day_labels[d] for d in day_order])
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill

        # Build weekly plan from grouped_data["tags"]
        table_data = {cat: {day: [] for day in day_order} for cat in categories}
        score_key = "score1" if concept_key == "concept1" else "score2"
        max_points = lesson_1["full"]["total_points"] if concept_key == "concept1" else lesson_2["full"]["total_points"]
        print("✅ Using max_points for", concept_key, ":", max_points)


        for tag in grouped_data.get("tags", []):
            score = tag.get(score_key)
            name = tag["name"]
            max_points = lesson_1["total_points"] if concept_key == "concept1" else lesson_2["total_points"]
            group = get_instruction_group(score, max_points)

            if group in table_data:
                for day in schedule_map.get(group, []):
                    table_data[group][day].append(name)

        for group_key, label in categories.items():
            row = [label]
            for day in day_order:
                names = ", ".join(table_data[group_key][day]) or "No group today for these student"
                row.append(names)
            ws.append(row)

        # Wrap text

        for row in ws.iter_rows(min_row=2, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Autofit columns (use helper)
        autofit_columns(ws)


        # Row fills by group
        for row in ws.iter_rows(min_row=ws.max_row - len(categories) + 1, max_col=6):
            group_label = row[0].value
            fill_color = group_colors.get(group_label)
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    return wb

#def generate_excel_view_V2(request):
    grouped_data = request.session.get("grouped_data", {})

    print("✅ grouped_data keys:", list(grouped_data.keys()))

    export_data = request.session.get("new_entries", [])
    lessons = request.session.get("lessons", {})

    print("DEBUG export_data:", export_data)
    print("DEBUG lessons:", lessons)

    wb = generate_excel(export_data, lessons)  # must return a Workbook

    # Write workbook into a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Sort2Support_{datetime.now():%Y%m%d}.xlsx"'
    return response

#def generate_excel_V2(export_data, lessons):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Main Sheet ---
    today_str = datetime.now().strftime("%Y-%m-%d")
    main_title = f"{today_str} Weekly Groupings"
    ws_main = wb.active
    ws_main.title = normalize_sheet_name(main_title, with_date=False)

    headers = ["Name", "Concept", "Score"]
    ws_main.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws_main[1]:
        cell.font = header_font
        cell.fill = header_fill

    for name, concept, score in export_data:
        ws_main.append([name, concept, score])
        score_cell = ws_main.cell(row=ws_main.max_row, column=3)
        fill = get_fill(score)
        if fill:
            score_cell.fill = fill

    ws_main.freeze_panes = "A2"
    print("✅ Lessons keys:", list(lessons.keys()))

    # --- Lesson Sheets ---
    for lesson_title, students in lessons.items():
        sheet_title = normalize_sheet_name(lesson_title)
        if sheet_title not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_title)
            ws.append(["Student", "Score"])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        else:
            ws = wb[sheet_title]

        for name, score in students:
            ws.append([name, score])
            row_idx = ws.max_row
            fill = get_fill(score)
            if fill:
                ws.cell(row=row_idx, column=1).fill = fill
                ws.cell(row=row_idx, column=2).fill = fill

        ws.freeze_panes = "A2"

    # ✅ Return the workbook, not a response
    return wb

# --- BACKUP: original export_grouped_excel ---
# def export_grouped_excel(request):
#     grouped_data = request.session.get("grouped_data")
#     lesson_1 = request.session.get("lesson_1")
#     lesson_2 = request.session.get("lesson_2")
#     ...
#        # Step 3: Preview table → update preview_data from rendered inputs
#        if "save_roster" in request.POST:
#            student_count = int(request.POST.get("student_count", 0))
#            preview_data = []
#            for i in range(1, student_count + 1):
#                name = request.POST.get(f"name_{i}", "").strip()
#                score1 = request.POST.get(f"score1_{i}", "")
#                score2 = request.POST.get(f"score2_{i}", "")
#                try:
#                    score1 = int(score1) if score1 != "" else 0
#                except ValueError:
#                    score1 = 0
#                try:
#                    score2 = int(score2) if score2 != "" else 0
#                except ValueError:
#                    score2 = 0
#                if name:
#                    preview_data.append({
#                        "name": name,
#                        "score1": score1,
#                        "score2": score2,
#                    })
#            request.session["preview_data"] = preview_data
#            request.session["entry_mode"] = "preview"
#            messages.success(request, "✅ Roster updated from preview.")

#        # Step 4: Upload roster → preview only
#        if "upload" in request.POST:
#            file = request.FILES.get("file") or request.FILES.get("roster")
#            if file:
#                print("✅ File received:", file.name)
#                wb = openpyxl.load_workbook(file)
#                sheet = wb.active
#                preview_data = []
#                for row in sheet.iter_rows(min_row=2, values_only=True):
#                    name, score1, score2 = row[:3]
#                    if name:
#                        preview_data.append({
#                            "name": name.strip(),
#                            "score1": score1 or 0,
#                            "score2": score2 or 0,
#                        })
#                request.session["preview_data"] = preview_data
#                request.session["entry_mode"] = "preview"
#                preview_mode = True
#                messages.success(request, "✅ Roster uploaded for preview.")
#                print("✅ Preview data in session:", preview_data)

#                Roster.objects.create(
#                    user=request.user,
#                    name="Latest Roster",
#                    data=preview_data
#                )
