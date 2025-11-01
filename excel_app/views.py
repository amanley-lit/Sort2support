from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import StudentFormSet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from main.models import Student
from .utils import parse_excel  # generate_excel - add later?
from django.views.decorators.http import require_POST
from django import forms
from django.utils.text import slugify
from django.urls import reverse
from django.forms import modelformset_factory
import json
from django.conf import settings
import os
import pandas as pd
from .utils.grouping import assign_group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse



from django.shortcuts import redirect
from excel_app.utils.parse_excel import parse_excel # or parse_excel
from .utils.grouping import assign_group 

@login_required
def saved_roster(request):
    lessons = load_ufli_lessons()
    students = request.session.get("students", [])
    has_saved_roster = bool(students)

    test_scores = {"score 1": 88, "score 2": 92}  # lowercase keys to match your template

    return render(request, "main/dashboard.html", {
        "lessons": lessons,
        "students": students,
        "has_saved_roster": has_saved_roster,
        "test_scores": test_scores,  # ✅ include this for filter testing
    })

class AddStudentForm(forms.Form):
    name = forms.CharField(max_length=100)
    ufli_score_1 = forms.IntegerField(required=False)
    ufli_score_2 = forms.IntegerField(required=False)

StudentFormSet = modelformset_factory(
    Student,
    fields=["name", "ufli_score_1", "ufli_score_2"],
    extra=5,  # Start with 5 rows
    can_delete=True
)


@login_required
def reset_session(request):
    keys_to_clear = [
        "uploaded_students",
        "selected_concepts",
        "saved_students",
        "file_uploaded",
    ]
    for key in keys_to_clear:
        request.session.pop(key, None)

    return redirect("main:sort2support")


@login_required
def add_student(request):
    if request.method == "POST":
        form = AddStudentForm(request.POST)
        if form.is_valid():
            Student.objects.create(
                name=form.cleaned_data["name"],
                ufli_score_1=form.cleaned_data["ufli_score_1"],
                ufli_score_2=form.cleaned_data["ufli_score_2"],
                teacher=request.user
            )
            messages.success(request, "✅ Student added.")
            url = reverse("excel_app:preview_uploaded_students") + "?edit=true"
            return redirect(url)

    else:
        form = AddStudentForm()

    return render(request, "excel_app/add_student.html", {"form": form})
@require_POST
@login_required
def delete_student(request, student_id):
    student = Student.objects.filter(id=student_id, teacher=request.user).first()
    if student:
        student.delete()
        messages.success(request, f"🗑️ Deleted {student.name}.")
    else:
        messages.error(request, "Student not found or unauthorized.")
    url = reverse("excel_app:preview_uploaded_students") + "?edit=true"
    return redirect(url)


@login_required
def manual_roster_entry(request):
    if request.method == "POST":
        formset = StudentFormSet(request.POST)
        if formset.is_valid():
            for form in formset:
                name = form.cleaned_data.get("name")
                score_1 = form.cleaned_data.get("ufli_score_1")
                score_2 = form.cleaned_data.get("ufli_score_2")
                if name:
                    Student.objects.create(
                        teacher=request.user,
                        name=name,
                        ufli_score_1=score_1,
                        ufli_score_2=score_2,
                    )
            messages.success(request, "✅ Students added successfully!")
            return redirect("main:dashboard")
    else:
        formset = StudentFormSet()

    return render(request, "excel_app/manual_entry.html", {"formset": formset})

# --- Excel parsing helper ---

def parse_excel(file):
    """Read Excel file and return list of dicts (one per row)."""
    df = pd.read_excel(file)
    return df.to_dict(orient="records")


def download_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Styled header row
    headers = ["Name", "Score 1", "Score 2"]
    ws.append(headers)

    header_font = Font(name="Comic Sans MS", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Sort2Support_Sample_Template.xlsx"'
    wb.save(response)
    return response



@login_required
def parse_excel_upload(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("roster") or request.FILES.get("file")
        if uploaded_file:
            request.session.pop("uploaded_students", None)
            request.session.pop("saved_students", None)
            request.session["file_uploaded"] = True

            rows = parse_excel(uploaded_file)

            def normalize(header):
                return header.strip().lower().replace("_", " ")

            rows = [{normalize(k): v for k, v in row.items()} for row in rows]

            score_columns = ["score 1", "score 2"]
            score_keys = ["ufli_score_1", "ufli_score_2"]

            preview_data = []
            merged = {}

            for row in rows:
                name = row.get("name", "").strip()
                if not name:
                    continue

                student_data = {
                    "name": name,
                    "ufli_score_1": "",
                    "ufli_score_2": "",
                }

                preview_row = {
                    "name": name,
                    "score1": "",
                    "score2": "",
                }

                for col, key in zip(score_columns, score_keys):
                    raw_score = row.get(col, "")
                    try:
                        score = int(float(raw_score)) if raw_score not in ["", None] else None
                    except (ValueError, TypeError):
                        score = None

                    student_data[key] = score
                    preview_row[f"score{col[-1]}"] = score

                merged[name] = student_data
                preview_data.append(preview_row)

            request.session["preview_data"] = preview_data
            request.session["uploaded_students"] = list(merged.values())
            request.session["score_columns"] = score_columns
            request.session["score_keys"] = score_keys

            messages.success(request, "Student roster uploaded successfully!")
            return redirect("main:dashboard")

        messages.error(request, "⚠️ No file was uploaded. Please select a file and try again.")
        return redirect("main:dashboard")

    return redirect("main:dashboard")




@login_required
def save_student_scores(request):
    if request.method != "POST":
        return HttpResponse("No data to save.")

    is_edit_mode = request.GET.get("edit") == "true"

    if is_edit_mode:
        saved_count = 0

        student_ids = [s.id for s in Student.objects.filter(teacher=request.user)]
        for student_id in student_ids:
            student = Student.objects.get(id=student_id)

            score_1_raw = request.POST.get(f"ufli_score_1_{student_id}", "").strip()
            score_2_raw = request.POST.get(f"ufli_score_2_{student_id}", "").strip()
            
            student.ufli_score_1 = int(score_1_raw) if score_1_raw.isdigit() else None
            student.ufli_score_2 = int(score_2_raw) if score_2_raw.isdigit() else None

            print(f"✅ Saving {student.name}: {student.ufli_score_1}, {student.ufli_score_2}")
            student.save()
            saved_count += 1

        messages.success(
            request,
            f"✅ Saved {saved_count} students."
        )
        return redirect(reverse("main:dashboard") + "#sort2support")



    # Upload mode: save new students
    saved_count = 0

    total_rows = int(request.POST.get("total_rows", 0))

    for i in range(1, total_rows + 1):
        name = request.POST.get(f"name_{i}", "").strip()
        ufli_score_1_raw = request.POST.get(f"ufli_score_1_{i}", "").strip()
        ufli_score_2_raw = request.POST.get(f"ufli_score_2_{i}", "").strip()

        ufli_score_1 = int(ufli_score_1_raw) if ufli_score_1_raw.isdigit() else None
        ufli_score_2 = int(ufli_score_2_raw) if ufli_score_2_raw.isdigit() else None


        Student.objects.create(
            name=name,
            ufli_score_1=ufli_score_1,
            ufli_score_2=ufli_score_2,
            teacher=request.user

        )
        saved_count += 1

    request.session.pop("uploaded_students", None)
    messages.success(request, f"✅ Saved {saved_count} students.")
    return redirect("main:dashboard")

@login_required
def preview_uploaded_students(request):
    is_edit_mode = request.GET.get("edit") == "true"

    if is_edit_mode:
        students = Student.objects.filter(teacher=request.user)
        table = convert_students_to_dicts(students)
    else:
        table = request.session.get("uploaded_students", [])

    return render(request, "excel_app/preview.html", {
        "table": table,
        "is_edit_mode": is_edit_mode,
    })
def add_group_color_highlighting(ws, start_row=2, last_col="E", group_col="B"):
    """
    Highlight entire rows based on whether the group label in group_col
    contains 'Red', 'Yellow', 'Green', or 'Blue'.
    The range automatically extends to ws.max_row.
    """

    end_row = ws.max_row  # dynamically detect last row with data

    colors = {
        "Red":    "F4CCCC",  # light red
        "Yellow": "FFF2CC",  # light yellow
        "Green":  "D9EAD3",  # light green
        "Blue":   "CFE2F3",  # light blue
    }

    for keyword, hex_color in colors.items():
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        # Formula: look for the keyword anywhere in the group label column
        formula = f'ISNUMBER(SEARCH("{keyword}",${group_col}{start_row}))'
        ws.conditional_formatting.add(
            f"A{start_row}:{last_col}{end_row}",
            FormulaRule(formula=[formula], fill=fill)
        )
@login_required
def edit_uploaded_score(request):
    data = request.session.get("new_entries", [])

    if request.method == "POST":
        formset = StudentFormSet(request.POST)
        if formset.is_valid():
            new_data = []
            for i, form in enumerate(formset):
                name = data[i][0] if i < len(data) else form.cleaned_data.get("name")
                score_1 = form.cleaned_data.get("ufli_score_1")
                score_2 = form.cleaned_data.get("ufli_score_2")
                if name:
                    new_data.append([name, score_1, score_2])
            request.session["new_entries"] = new_data
            data = new_data
    else:
        formset = StudentFormSet(
            initial=[
                {"name": entry[0], "ufli_score_1": entry[1], "ufli_score_2": entry[2]}
                for entry in data
            ]
        )

    for i, form in enumerate(formset):
        if i < len(data) and data[i][0]:
            form.fields["name"].widget.attrs["readonly"] = True

 
    return render(request, "excel_app/upload.html", {"formset": formset, "data": data})
@login_required
def clear_preview_scores(request):
    data = request.session.get("new_entries", [])
    cleared_data = [[entry[0], None, None] for entry in data if entry]
    request.session["new_entries"] = cleared_data
    return redirect("main:upload_page")
# ---------- Load UFLI lessons ----------
def load_ufli_lessons():
    path = os.path.join(settings.BASE_DIR, "main", "static", "main", "data", "ufli_lessons.json")
    with open(path, encoding="utf-8") as f:
        lessons = json.load(f)

    # Ensure every lesson has a total_points field
    for lesson in lessons:
        if "total_points" not in lesson:
            lesson["total_points"] = 5  # fallback

    return lessons

ufli_lessons = load_ufli_lessons()

@login_required
def save_students(request):
    if request.method == "POST":
        students = request.session.get("uploaded_students", [])
        teacher = request.user  # ✅ Get the logged-in teacher

        for student in students:
            Student.objects.update_or_create(
                name=student["name"],
                teacher=teacher,  # ✅ Link to teacher
                defaults={
                    "ufli_score_1": student["ufli_score_1"],
                    "ufli_score_2": student["ufli_score_2"],
                }
            )

        # ✅ Clear session before redirect
        request.session.pop("uploaded_students", None)
        request.session["file_uploaded"] = False

        messages.success(request, "✅ Students saved successfully! Scroll down to see your table!")
        return redirect("main:dashboard")

    messages.error(request, "⚠️ Save failed. Please try again.")
    return redirect("main:sort2support")


def parse_excel_upload(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        student_data = parse_excel(file)  # returns list of dicts with name, score1, score2

        # Save to session for grouping
        request.session["student_data"] = student_data

        return redirect("main:sort2support")  # jump straight to grouping preview

    return redirect("main:dashboard")




























































