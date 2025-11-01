import io
import pytest
from openpyxl import load_workbook
from excel_app.utils.export_excel import generate_excel


# Hex RGB values that openpyxl stores for fills
RED = "00FFC7CE"     # light red
YELLOW = "00FFEB9C"  # light yellow
GREEN = "00C6EFCE"   # light green
BLUE = "009CC3E6"    # light blue

@pytest.mark.django_db
def test_generate_excel_creates_three_sheets_and_headers():
    export_data = [
        ["Alice", 45, 88],   # red + green
        ["Bob", 72, 100],    # green + blue
        ["Charlie", 55, 65], # yellow + yellow
    ]

    response = generate_excel(export_data)
    wb = load_workbook(io.BytesIO(response.content))

    # Check sheet names
    assert wb.sheetnames == ["Data", "Summary", "Teacher View"]

    # Check headers in Data sheet
    ws1 = wb["Data"]
    headers = [cell.value for cell in ws1[1]]
    assert headers == ["Name", "Score C1", "Score C2"]

    # Check Summary sheet has totals
    ws2 = wb["Summary"]
    assert ws2["A1"].value == "Total Students"
    assert ws2["B1"].value == 3

    # Check Teacher View numbering
    ws3 = wb["Teacher View"]
    assert ws3["A2"].value == 1
    assert ws3["B2"].value == "Alice"


@pytest.mark.django_db
def test_generate_excel_highlighting_colors():
    export_data = [
        ["Low", 40, None],     # should be red
        ["Borderline", 55, None], # should be yellow
        ["Good", 75, None],    # should be green
        ["Perfect", 100, None] # should be blue
    ]

    response = generate_excel(export_data)
    wb = load_workbook(io.BytesIO(response.content))
    ws1 = wb["Data"]

    # Row 2: 40 -> red
    assert ws1["B2"].fill.start_color.rgb == RED

    # Row 3: 55 -> yellow
    assert ws1["B3"].fill.start_color.rgb == YELLOW

    # Row 4: 75 -> green
    assert ws1["B4"].fill.start_color.rgb == GREEN

    # Row 5: 100 -> blue
    assert ws1["B5"].fill.start_color.rgb == BLUE
