# excel_app/utils/export_excel.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.http import HttpResponse


def generate_excel(export_data):
    wb = Workbook()

    # --- Sheet 1: Raw Data ---
    ws1 = wb.active
    ws1.title = "Data"

    headers = ["Name", "Score C1", "Score C2"]
    ws1.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Write rows with conditional highlighting
    for row in ws1.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if cell.value is not None:
                if cell.value < 50:
                    # Red fill
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif 50 <= cell.value < 70:
                    # Yellow fill
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif 70 <= cell.value < 100:
                    # Green fill
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value == 100:
                    # Blue fill
                    cell.fill = PatternFill(start_color="9CC3E6", end_color="9CC3E6", fill_type="solid")


    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet(title="Summary")
    ws2.append(["Total Students", len(export_data)])
    avg_c1 = sum([r[1] for r in export_data if r[1] is not None]) / max(1, len(export_data))
    avg_c2 = sum([r[2] for r in export_data if r[2] is not None]) / max(1, len(export_data))
    ws2.append(["Average Score C1", avg_c1])
    ws2.append(["Average Score C2", avg_c2])

    # Auto-adjust column widths
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Teacher View ---
    ws3 = wb.create_sheet(title="Teacher View")
    ws3.append(["#", "Student", "C1", "C2"])
    for i, row in enumerate(export_data, start=1):
        ws3.append([i, row[0], row[1], row[2]])

    # Style headers
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Auto-adjust column widths
    for col in ws3.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws3.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row
    ws3.freeze_panes = "A2"

    # --- Return response ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"student_export_{timestamp}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response