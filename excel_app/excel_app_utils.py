from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ---------- Export: Grouped Workbook ----------
def build_grouped_workbook(grouped_data):
    """Generates an Excel workbook from grouped data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Students"

    ws.append(["Group", "Name", "Score"])
    for group, students in grouped_data["concept1"].items():
        for name, score in students:
            ws.append([group, name, score])

    ws.append([])
    ws.append(["Group", "Name", "Score"])
    for group, students in grouped_data["concept2"].items():
        for name, score in students:
            ws.append([group, name, score])

    return wb

# ---------- Export: Class Template ----------
def build_class_template():
    """Generates a blank class template workbook for teacher input."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Template"

    ws.append(["Name", "score 1", "score 2"])
    for i in range(1, 21):
        ws.append([f"Student {i}", "", ""])

    return wb

# ---------- Export: Raw Student Data ----------
def build_students_workbook(students):
    """Generates a workbook from raw student data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Scores"

    ws.append(["Name", "Score 1", "Score 2"])
    for student in students:
        ws.append([student.name, student.score_1, student.score_2])

    return wb
